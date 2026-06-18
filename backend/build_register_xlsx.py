# Builds the Rusiyaa Google Sheets database templates.
#
# Produces TWO files:
#   rusiyaa-register.xlsx      -> full workbook, both tabs corrected
#                                 (use this if you want to start fresh / "Replace spreadsheet")
#   rusiyaa-batches-tab.xlsx   -> ONLY the Batches register tab
#                                 (use this to "Insert new sheet(s)" without losing data)
#
# Tab roles:
#   "Batches"          (Table 1, PUBLIC trace) — col A MUST be the Batch Number lookup key
#   "Admin Dashboard"  (Table 2, ADMIN ONLY)   — procurement / packing / distribution
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

INK   = "1C0E07"   # clove brown (header fill)
CREAM = "F4E8D2"   # muslin cream (header text)

header_font  = Font(name="Calibri", bold=True, color=CREAM, size=11)
header_fill  = PatternFill("solid", fgColor=INK)
header_align = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="DDCBA8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_sheet(ws, headers, widths, rows):
    ws.append(headers)
    for c, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = (
            header_font, header_fill, header_align, border)
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def grams_dropdown(ws, col_letter):
    dv = DataValidation(type="list", formula1='"50,100"', allow_blank=True,
                        showDropDown=False)
    dv.error, dv.errorTitle = "Choose 50 or 100", "Grams"
    dv.prompt, dv.promptTitle = "Pick the pack size", "Grams"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}500")


# ----- Batches register (public lookup; col A = Batch Number) -----
batch_headers = ["Batch Number", "Product Name", "Grams", "SKU",
                 "Manufacturing Details", "FSSAI License No"]
batch_widths  = [16, 20, 8, 14, 64, 18]
batch_rows = [
    # RSY-2605-A12 is the sample shown on the website's home page — keep it.
    ["RSY-2605-A12", "Garam Masala",      100, "RSY-GM-100",
     "Mfg 30-May-2026 · Best before 29-May-2027 · Erode (Erode Spice Suppliers) · Mill Line 2 · QA: S. Meenakshi",
     "12426XXXXXX012"],
    ["RSY-2606-R12", "Red Chilli Powder",  50, "RSY-RC-050",
     "Mfg 04-Jun-2026 · Best before 03-Jun-2027 · Guntur (Guntur Agro Traders) · Mill Line 1 · QA: S. Meenakshi",
     "12426XXXXXX012"],
    ["RSY-2606-T18", "Turmeric Powder",   100, "RSY-TU-100",
     "Mfg 02-Jun-2026 · Best before 01-Jun-2027 · Salem (Salem Turmeric Co-op) · Mill Line 3 · QA: R. Karthik",
     "12426XXXXXX012"],
]

# ----- Admin Dashboard (your real procurement data, 11 columns) -----
admin_headers = ["Date", "Vendor Name", "Product", "Quantity", "Purchased Date",
                 "Invoice Number", "Packed On", "Grams per Unit", "No. of Units",
                 "Distributed Name", "Distributed Date"]
admin_widths  = [14, 22, 20, 12, 16, 16, 14, 14, 12, 24, 16]
admin_rows = [
    ["01-Jun-2026", "Guntur Agro Traders",  "Dry Red Chilli",    "250 kg",
     "01-Jun-2026", "INV-2381", "04-Jun-2026",  50, 1800, "Chennai South Distributors", "06-Jun-2026"],
    ["28-May-2026", "Salem Turmeric Co-op",  "Turmeric Fingers",  "400 kg",
     "28-May-2026", "INV-2374", "02-Jun-2026", 100, 3200, "Madurai Spice Mart",         "05-Jun-2026"],
    ["25-May-2026", "Erode Spice Suppliers", "Whole Garam Spices","120 kg",
     "25-May-2026", "INV-2369", "30-May-2026", 100, 1050, "Coimbatore Wholesale Hub",   "03-Jun-2026"],
]

# ---------- File 1: full corrected workbook ----------
wb = openpyxl.Workbook()
b = wb.active
b.title = "Batches"
style_sheet(b, batch_headers, batch_widths, batch_rows)
grams_dropdown(b, "C")

a = wb.create_sheet("Admin Dashboard")
style_sheet(a, admin_headers, admin_widths, admin_rows)
grams_dropdown(a, "H")   # Grams per Unit column
wb.save("rusiyaa-register.xlsx")

# ---------- File 2: just the Batches tab (insert without replacing) ----------
wb2 = openpyxl.Workbook()
b2 = wb2.active
b2.title = "Batches"
style_sheet(b2, batch_headers, batch_widths, batch_rows)
grams_dropdown(b2, "C")
wb2.save("rusiyaa-batches-tab.xlsx")

print("wrote rusiyaa-register.xlsx (Batches + Admin Dashboard)")
print("wrote rusiyaa-batches-tab.xlsx (Batches only)")
